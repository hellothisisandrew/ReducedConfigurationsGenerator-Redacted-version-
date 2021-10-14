from typing import NewType
import sheet
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, colors
from sheet import dict_error, compare, get_column_number, filterSearch


# Removed senstive strings below -------------------------------------------   
AllMonitor = "*-----*" 
USB = "*-----*"                 
arrow = " --> "             
Monoprice = "*-----*"
plus = " + "
left = " ("
right = ")"                                                 #These strings were for formmating
on = "on"
MDPtoMDP = '*-----*'
MDPtoDP = '*-----*'
MultiMonitor = "*-----*"
BlankAdapter = " BlankAdapter "
TB = "*-----*"
USBC = "*-----*"
HDMI = "*-----*"
DP = "*-----*"
# Removed senstive strings above ---------------------------------------------

class colorCol(object):
    def __init__(self,y,x):
        self.x = x
        self.y = y
    
    def getX(self):
        return self.x
    def getY(self):
        return self.y



class excel_instance(object):


    

    
    
    def __init__(self, Excel_File,instance_type):
        self.scenariosSheet = None
        self.resultsSheet = None
        self.issuesSheet = None

        self.firstTestColumn = 0    

        self.index = 10             
        self.issues = "*-----*" #Removed String

        

        if instance_type == AllMonitor:    
           self.scenarios = "*-----*" #Removed String
           self.results = "*-----*" #Removed String
           self.scenariosSheet = sheet.AllMonitor(Excel_File, self.scenarios)
        else:
            self.scenarios = "*-----*"  #Removed String
            self.results = "*-----*"    #Removed String
            self.scenariosSheet = sheet.USB(Excel_File, self.scenarios)  
            
        
        self.issuesSheet = sheet.Standard(Excel_File,self.issues) 
        self.resultsSheet = sheet.Standard(Excel_File, self.results, self.index)
        

    
    def generateRowArray(self, row):
        array = self.scenariosSheet.generateRowList(row) 
        return array
    '''         
    
    def searchForInfo(self, info, array):
        info = str(info).lower().replace(" ", "") 
        for col,x in enumerate(array):              
            if str(x).lower().replace() == info:    
                return get_column_letter(col)                          

        return dict_error                           
    '''

   

class display_instance(excel_instance):
    
    def __init__(self,Excel_File):
        excel_instance.__init__(self,Excel_File, AllMonitor)
        self.Internal_Inventory = "*----*."   #Removed String

        

    
    
       
    
    
    def checkMonitorDISP(self,text):
        text = str(text).replace(" ", "").lower()
        if re.search("disp",text) == None:
            return False
        else:
            return True
    
    
    def getMon(self, array, number):
        if number == 1:
            col = self.scenariosSheet.getInfoColumn("*----*") #Removed string
        else:
            col = self.scenariosSheet.getInfoColumn("*----*") #Removed string
        
        toCheck =  str(array[get_column_number(col) - 1]).replace("(DP)","").replace("(HDMI)","").replace("(USB-C)","").replace("(TB)", "") 

        
        if len(toCheck) < 5:        
            return None
        else:
            return toCheck

    
    
    def getCon(self, array, number):
        if number == 1:
             col = self.scenariosSheet.getInfoColumn("*-----*")  #Removed string
        else:
            col = self.scenariosSheet.getInfoColumn("*-----*")   #Removed string
        
        toCheck =  array[get_column_number(col) - 1]
        if len(toCheck) < 2:
            return None
        else:
            return toCheck
    
        
    
    def getDUT(self,array):
        col = self.scenariosSheet.getInfoColumn("*-----*") #Removed string

        processor = self.scenariosSheet.getInfoColumn("*-----*")  #Removed string

        if processor == dict_error:
            return array[get_column_number(col) - 1]
        else:
            return array[get_column_number(col) - 1] + "  " + array[get_column_number(processor) - 1] 
        
    
    def getAdp(self, array):
        col = self.scenariosSheet.getInfoColumn("*-----*")  #Removed string

        return array[get_column_number(col) - 1]
    
    def getDock(self,array, dock = "*-----*"):  #Removed string     
        if dock == "*-----*":   #Removed string
            col = self.scenariosSheet.getInfoColumn("*-----*")  #Removed string
        elif dock == "*-----*": #Removed string
            col = self.scenariosSheet.getInfoColumn("*-----*")  #Removed string
        else:
            col = self.scenariosSheet.getInfoColumn("*-----*")  #Removed string
        

        return array[get_column_number(col) - 1]

    
    
    
    

    
    
    def simpleConfig(self,DUT,monitor1, adapter, connection1, monitor2, connection2):    
        if connection1 != None:
            conn1 = left + str(connection1) + right
        else:
            conn1 = ""
        if connection2 != None:
            conn2 = left + str(connection2) + right
        
        
        if compare(adapter, "on") or compare(adapter,"-"): 
            if compare(connection1,"MDP"):
                adapter = MDPtoMDP
            else:
                adapter = MDPtoDP

        
        if compare(connection1, "DP"):  #Removed string
            cable = Monoprice
        else: 
            cable = "*-----*"  #Removed string

        if monitor2 == None:                
            return DUT + arrow + adapter + arrow + monitor1 + conn1        
        else:
            if self.checkMonitorDISP(monitor1):
                return DUT + arrow + adapter + arrow + monitor1 + conn1 + plus + adapter + arrow + monitor2 + conn2
            else:
                return DUT + arrow + adapter + arrow + monitor1 + conn1 + arrow + cable + arrow+ monitor2 + conn2

    
    
    def dockConfig(self, DUT, Monitor1, connection1, dock , MM):       
        
        connection1 = str(connection1).replace(" ","").upper()  
        if MM == True:

            if filterSearch("*-----*", str(dock)): #Removed string
                if connection1 != USBC:
                    connection2 = left + USBC + right
                else: 
                    connection2 = left + DP + right
            elif filterSearch("*-----*",str(dock)) or  filterSearch("*-----*",str(dock)): #Removed string
                connection2 = left + DP + right
            elif filterSearch( "*-----*", str(dock)) or filterSearch("*-----*", str(dock)) or filterSearch("*-----*",str(dock)) or filterSearch("*-----*", str(dock)): #Removed string
                connection2 = left + USBC + right
            elif filterSearch("*-----*",str(dock)): #Removed string
                if connection1 == HDMI:
                    connection2 = left + DP + right
                else:
                    connection2 = left + HDMI + right
            elif filterSearch("*-----*",str(dock)) or filterSearch("*-----*",str(dock)) or filterSearch("*-----*",str(dock)):  #Removed string
                if connection1 == USBC or connection1 == TB:
                    connection2 = left + DP + right
                else:
                    connection2 = left + USBC + right
            else:
                connection2 = left + HDMI + right

            return str(DUT) + arrow + str(dock) + arrow + BlankAdapter + arrow + str(Monitor1) + left +str(connection1) + right + plus + BlankAdapter + arrow + MultiMonitor + connection2
        else:       
            return DUT + arrow + dock + arrow + BlankAdapter + arrow + Monitor1 + left + connection1 + right
            



    def configGen(self):
            firstRow = self.scenariosSheet.firstRow()     
            lastRow = self.scenariosSheet.maxRow()
            colorColumns = []   

            while(firstRow <= lastRow):

                array = self.generateRowArray(firstRow)

                issues = issueList(self.issuesSheet)            
                
                for index, value in enumerate(array):
                    if self.scenariosSheet.getTestName(get_column_letter(index + 1)) != dict_error: 
                        if issues.correctFormat(value) == False:
                            
                            color = colorCol(firstRow, index + 1)
                            colorColumns.append(color)
                    
                    #Code is somewhat redundent I could have cleaned it up
                    if re.search("*-----*", value): 
                        if re.search("*-----*",self.scenariosSheet.getTestName(index + 1)): 
                            if re.search("*-----*]",self.scenariosSheet.col_to_title[get_column_letter(index + 1)]):
                               toPrint = self.dockConfig(self.getDUT(array),self.getMon(array,1),self.getCon(array,1), self.getDock(array),True) 
                               issues.addConfig(toPrint, value, "*-----*", True) 
                            else:
                                toPrint = self.dockConfig(self.getDUT(array),self.getMon(array,1),self.getCon(array,1), self.getDock(array),False) 
                                issues.addConfig(toPrint, value, "*-----*") 

                        elif re.search("*-----*",self.scenariosSheet.getTestName(index + 1)): 
                            if re.search("*-----*",self.scenariosSheet.col_to_title[get_column_letter(index + 1)]):
                                toPrint = self.dockConfig(self.getDUT(array),self.getMon(array,1),self.getCon(array,1), self.getDock(array,"*-----*"),True) 
                                issues.addConfig(toPrint, value, "*-----*", True) 
                            else:
                                toPrint = self.dockConfig(self.getDUT(array),self.getMon(array,1),self.getCon(array,1), self.getDock(array,"*-----*"),False) 
                                issues.addConfig(toPrint, value, "*-----*") 

                        elif re.search("*-----*",self.scenariosSheet.getTestName(index + 1)): 
                            if re.search("*-----*",self.scenariosSheet.col_to_title[get_column_letter(index + 1)]):
                                toPrint = self.dockConfig(self.getDUT(array),self.getMon(array,1),self.getCon(array,1), self.getDock(array, "*-----*"),True) 
                                issues.addConfig(toPrint, value, "*-----*", True) 
                            else:
                                toPrint = self.dockConfig(self.getDUT(array),self.getMon(array,1),self.getCon(array,1), self.getDock(array,"*-----*"),False) 
                                issues.addConfig(toPrint, value, "*-----*") 

                        else:
                            toPrint = self.simpleConfig(self.getDUT(array),self.getMon(array,1),self.getAdp(array),self.getCon(array,1),self.getMon(array,2),self.getCon(array,2))                            
                            issues.addConfig(toPrint, value, "*-----*")
                           
                issues.output()         
                firstRow += 1

            issues.finishAndColor(self.scenarios,colorColumns) 



class issue(object):

    def __init__(self, name, issuesSheet):       
        self.issueName = str(name)
        self.configList = ['','','','','','',''] 

        self.issueSheet = issuesSheet

    
    
    
    def haveISeenIt(self, testName, MultiMonitor = False):
        if MultiMonitor == True:
            offset = 3
        else:
            offset = 0

        if re.search("*-----*",testName):
            if self.configList[0] != '':
                return True
            else:
                return False
        if re.search("*-----*", testName):
            if self.configList[1 + offset] != '':
                return True 
            else:
                return False
        if re.search("*-----*", testName):
             if self.configList[2 + offset] != '':
                 return True
             else:
                 return False
        if re.search("*-----*",testName):
            if self.configList[3 + offset] != '':
                return True
            else:
                return False
        print("*-----*")
        return False


    
    def newConfig(self,config, testName, MultiMonitor = False):
        if MultiMonitor == True:
            offset = 3
        else:
            offset = 0

        if re.search("*-----*",testName):
            self.configList[0] = config
        elif re.search("*-----*", testName):
            self.configList[1 + offset] = config
        elif re.search("*-----*", testName):
            self.configList[2+ offset] = config
        elif re.search("*-----*",testName):
            self.configList[3 + offset] = config
        else:
            print("*-----*")
    '''
    
    def textThisissues(self, text):
        if str(text.lower().replace(" ","")) == self.issueName.lower().replace(" ", ""): 
            return True
        else:
            return False
    ''' 
    
    def name(self):
        return self.issueName

    
    def isEmpty(self):
        for x in self.configList:
            if x != '':
                return False
        return True
            
    
    def write(self):
        
        BugID  = self.issueSheet.getColumnInfo("*-----*")
        UniversalID = self.issueSheet.getColumnInfo("*-----*")
        rowT = -1
        
        for index, value in enumerate(UniversalID):
            if compare(self.issueName, value) or compare(self.issueName, BugID[index]):
                rowT = index
                break
        if rowT == -1:
            print("*-----*")
        
        for value in self.configList:
            if value != '':
                self.issueSheet.writeOrAppend(rowT + 2,self.issueSheet.maxColumn() + 2, value) 



class issueList(object):
    
    def __init__(self, issuesSheet):              
        
        self.issuesSheet = issuesSheet
        self.issuesListSmoke  = []
        self.actualIssue = []
        
        self.generateList(self.issuesSheet)            
                                                                                                                
        self.createActualIssues()       


    def generateList(self, issuesSheet):
       BugId = issuesSheet.getColumnInfo("*-----*") 
       UniID = issuesSheet.getColumnInfo("*-----*") 
       
       
       
       for indx, val in enumerate(UniID):
           
           if val == 'None' and BugId[indx] == 'None':      
               continue					
           if val == 'None':                                
                toAdd = issue(BugId[indx],issuesSheet)
           else:
                toAdd = issue(val,issuesSheet)                          

           self.issuesListSmoke.append(toAdd) 
    
       return self.issuesListSmoke

    
    def createActualIssues(self):
        for index, item in enumerate(self.issuesListSmoke):
            newIssue = issue(item.name(),self.issuesSheet) 
            self.actualIssue.append(newIssue)

    
    def addConfig(self, config, fails, testName, MultiMonitor = False):
        
        
        for index, item in  enumerate(self.issuesListSmoke):
            if re.search(str(item.name()).replace(" ", "").replace("_","").lower(),str(fails).replace(" ", "").replace("_","").lower()): 
                
                if self.actualIssue[index].haveISeenIt(testName, MultiMonitor):
                    continue
                else:
                    self.actualIssue[index].newConfig(config,testName, MultiMonitor)

    def output(self):
        for item in self.actualIssue:
            if item.isEmpty():
                continue
            item.write()
     
 
    def finishAndColor(self,scenarios,colors):
        self.issuesSheet.saveAndClose(scenarios,colors)
    
    
    #This could have been broken into a few functions
    
    def correctFormat(self, testCell):
        testList = []                   
        testCell = str(testCell)
       
        if testCell == "Pass" or testCell == "N/A" or testCell == "on" or testCell == "None" or testCell == "-" or testCell == "": 
            return True
        
        if len(testCell) < 7 or testCell[0:7] != "Fail - ":
            return False
        
        if re.search("  ", str(testCell)):
            return False
        
        if testCell[len(testCell) - 1] == " ":
            return False
        
        if re.search("[a-z]", testCell[8:]):
            return False
        
        
        
        testList.append("")
        testNumber = 0      
        for index, char in enumerate(testCell[7:]):
            
            if char == " ":
                testNumber += 1
                testList.append("")
                continue                

            testList[testNumber] += char

        
        if len(testList) != len(set(testList)):
            return False

        
        for test in testList:
            isThere = False
            test = test.replace("_", " ")       

            for issue in self.issuesListSmoke: 
                if re.search("\.[1-4]", test):      
                    if issue.name() == test[:len(test) - 2]:
                        isThere = True
                        break
                else:
                    if issue.name() == test:
                        isThere = True
                        break

            if isThere == False:
                return False
        
        
        return True
