import openpyxl 
import xlwings 
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, colors, Alignment
from openpyxl.formatting import rule
from openpyxl.formatting.formatting import ConditionalFormattingList
import math








def get_column_number(column):
    column = str(column).lower()
    number = 0

    for indx, letter in enumerate(column):
        number += (ord(letter) - 96) * pow(26, indx)      
    return number

     



dict_error = "DOES NOT EXIST"         



def compare(string1,string2):
    if str(string1).lower().replace(" ","") == str(string2).lower().replace(" ",""):
        return True
    else:
        return False

def filterSearch(string1,string2):
    return re.search(str(string1).replace(" ","").lower(), str(string2).replace(" ", "").lower())
    






class Data(object):
   
    
    
    def __init__(self,excel_file,excel_sheet, index):
        self.col_to_title = {} 
        self.work_book =  openpyxl.load_workbook(excel_file)
        self.work_sheet = self.work_book[excel_sheet]
        self.row = index
        self.sheetName = excel_sheet
        self.fileName = excel_file
        self.max_column = self.work_sheet.max_column        
        self.__create_dic(index)

        
        self.writeworkbook = openpyxl.load_workbook(self.fileName, read_only = False)
        self.writeworksheet = self.writeworkbook[self.sheetName]

    
    
    def __create_dic(self, row_index):
        for row in self.work_sheet[row_index]:
            col_let = get_column_letter(row.column) 

            Is_None =  str(row.internal_value) 
            if Is_None != 'None':               
                self.col_to_title[col_let] = Is_None

    
    def detect_merged_cell(self, cell):
        if type(cell).__name__ == 'MergedCell':
            return True
        else:
            return False
    '''
    
    def column_title(self,text):
        try:
            to_return = self.col_to_title[text]
            return to_return
        except KeyError:
            return dict_error
    '''
    
    def Column_List(self):
        return self.col_to_title.values()
    
    
    def Title_Row(self):
        return self.row
    
    
    def generateRowList(self,row):
        toIterate = self.work_sheet.iter_cols( None, self.maxColumn(), row,row, True)
        toReturn = []
        for cell in toIterate:                  
            toReturn.append(str(cell)[2:len(cell) - 4])         
        return toReturn
    
    def maxColumn(self):
        return self.max_column

    
    def firstRow(self):
        return self.row + 1

    
    def maxRow(self):
        return self.work_sheet.max_row
    

    
    def writeOrAppend(self,rowT,columnT,text):
        current = self.writeworksheet.cell(row = rowT, column = columnT).value
        if current != None:
            current = current + "\n" +  str(text)
        else:
            current = text
        
        self.writeworksheet.cell(row = rowT, column = columnT).value = current
        self.writeworksheet.cell(row = rowT , column = columnT).alignment = Alignment(wrap_text=True)

    
    def colorCell(self,rowT,columnT, color):
        self.writeworksheet.cell(row = rowT, column = columnT).fill = PatternFill(start_color = color,fill_type = "solid")
        

    
    def saveAndClose(self,scenarios,colors):
        self.writeworkbook[scenarios].conditional_formatting = ConditionalFormattingList() 
        
        for colorC in colors:
            self.writeworkbook[scenarios].cell(row = colorC.getY(), column = colorC.getX()).fill = PatternFill(start_color = "FFFF0000",fill_type = "solid") 
                   
        self.writeworkbook.save( "sheetChecked----" + str(self.fileName).replace(".xlsm",".xlsx"))
        self.writeworkbook.close()








class AMUSB(Data):
    
    
    def __init__(self,excel_file, excel_sheet, Title_Index):
        self.col_to_test = {} 
        self.info_to_col = {} 
        self.Title_Index = Title_Index
        Data.__init__(self,excel_file,excel_sheet, Title_Index)
        self.__buildDicts()
       
    
    
    
    def __isTest(self, text):
       text = text.replace(" ", "")
       if None == re.search("[a|A][0-9]\.[0-9]*", text): 
           return False                                                                                                 
       else:                                                                                                            
           return True    
    
    
    
    

    
    def __buildDicts(self):
        for row in self.work_sheet[self.Title_Index - 1]:

            if(self.detect_merged_cell(row)):
                print("")
            elif self.__isTest(str(row.internal_value)):                                 
                self.col_to_test[str(get_column_letter(row.column))] = str(row.internal_value).replace(" ", "") 
            else:                                                                   
                info_txt = get_column_letter(row.column) + str(self.Title_Index)         
                self.info_to_col[str(self.work_sheet[info_txt].internal_value)] = get_column_letter(row.column)
       
    

    
    
    
    def getTestName(self,text):
        if isinstance(text, int):           
            col = get_column_letter(text)
        else:
            col = text
        col = col.upper().replace(" ","") 
        try:
            to_return = self.col_to_test[col]
            return to_return
        except KeyError:
            return dict_error
    
    
    
    def getInfoColumn(self, text):
        try:
            to_return = self.info_to_col[text]
            return to_return
        except KeyError:
            return dict_error
    '''
    
    def getInfoList(self):
        return self.info_to_col.keys()

    
    def getTestList(self):
        return self.col_to_test.values()
    '''






class AllMonitor(AMUSB):
    def __init__(self, excel_file, excel_sheet):    
        title_index = 3                                         
        AMUSB.__init__(self, excel_file,excel_sheet, title_index)
'''

class USB(AMUSB):
    def __init__(self,excel_file, excel_sheet):
        title_index = 9                                          
        AMUSB.__init__(self, excel_file, excel_sheet, title_index)
'''








class Standard(Data):
    def __init__(self, excel_file, excel_sheet, Title_Index = 1):
        Data.__init__(self,excel_file,excel_sheet,Title_Index)

        self.title_to_col =  {v: k for k, v in self.col_to_title.items()}     

    
    def getTitleColumn(self, text):
        try:
            to_return = self.title_to_col[text]
            return to_return
        except KeyError:
            return dict_error 

    
    
    
    def getColumnInfo(self, column_name):
        toReturn = []
        col = get_column_number(self.title_to_col[column_name])        
        array = list(self.work_sheet.iter_cols(min_col=col, max_col=col, min_row=self.firstRow(), max_row=None, values_only=True))[0]

        Not_none_values = [str(i or 'None') for i in array]     
        toReturn = list(Not_none_values)

        return toReturn  








class refLegends(object):
    
    def __init__(self,excel_file, excel_sheet):         
        self.wb = openpyxl.load_workbook(excel_file, data_only = True) 
        self.ws = self.wb[excel_sheet]              
        
        
        self.platformTitle = []
        self.tbDockTitle = []
        self.dockTitle = []
        self.cableTitle = []
        self.adapterTitle = []

        self.setLists() 
    
    
    
    
    def setLists(self):
        self.platformTitle =  self.titleFinder("*-----*") #Removed String
        self.tbDockTitle= self.titleFinder("*-----*")  #Removed String      
        self.dockTitle = self.titleFinder("*-----*") #Removed String
        self.cableTitle = self.titleFinder("*-----*")    #Removed String
        self.adapterTitle = self.titleFinder("*-----*")    #Removed String


    
    def titleFinder(self,searchString):
        for i in range(1, self.ws.max_row + 1):
            for j in range(1, self.ws.max_column + 1):
                if str(searchString).lower().replace(" ", "") == str(self.ws.cell(i,j).value).lower().replace(" ", ""): 
                    return i, j    

    
    
    def getListItemRC(self, itemName, tileRowCol):
        row = tileRowCol[0] + 1
        col = tileRowCol[1]

        while(row <= self.ws.max_row):
            if self.ws.cell(row,col).value == None: 
                return -1,-1

            if compare(self.ws.cell(row,col).value, itemName):            
                return row, col
            else:
                row += 1

        return -1,-1            
        



    def dockInfo(self, dockName, TB = False):
        dockAttribute = 0 
        if TB == False:
            rowCol = self.getListItemRC(dockName,self.dockTitle)
            dockAttribute = 5
        else:
            rowCol = self.getListItemRC(dockName, self.tbDockTitle)
            dockAttribute = 6
            

        toReturn = []

        if rowCol[0] == -1:         
            return dict_error

        for i in range(rowCol[1] + 1, rowCol[1] + dockAttribute):
            toReturn.append(self.ws.cell(rowCol[0],i).value)

        return toReturn

    
    def platformInfo(self, platform):
        rowCol = self.getListItemRC(platform,self.platformTitle)

        toReturn = []

        if rowCol[0] == -1:         
            return dict_error

        for i in range(rowCol[1] + 1, rowCol[1] + 3):
            toReturn.append(self.ws.cell(rowCol[0],i).value)

        return toReturn

 
    
    def cableInfo(self, cable):
        rowCol = self.getListItemRC(cable,self.cableTitle)

        toReturn = []

        if rowCol[0] == -1:         
            return dict_error

        for i in range(rowCol[1] + 1, rowCol[1] + 2):
            toReturn.append(self.ws.cell(rowCol[0],i).value)

        return toReturn



    
    def adapterInfo(self, adapter):
        pass

    def test1(self):
        while(True):
            listINq = input("List you are searching:::: ")
            print("\n")
            item = input("item you are looking for:::: ")
            print("\n")

            rowCol = self.getListItemRC(item, self.titleFinder(listINq))
            print("row: " + str(rowCol[0]) + "  col: " + str(rowCol[1]))
            print("\n")

    def test2(self):
        while(True):
            x = input("Press d for dock c for cable p for platform and tb for thunderbolt::: ")
            print("\n\n")
            if x == 'd':
                toPrint = self.dockInfo(input("*-----*")) #removed string
            elif x == 'c':
                toPrint = self.cableInfo(input("*------* ")) #Removed String
            elif x == 'p':
                toPrint = self.platformInfo(input("*------*")) #Removed String
            elif x == 't':
                toPrint = self.dockInfo(input("*---------* "), True) #Removed String
            else:
                print("try again")
                continue

            print("\n\n")
            print("-"*10 + "  dock info:-> \n")
            print(toPrint)
            print("\n\n")

            



